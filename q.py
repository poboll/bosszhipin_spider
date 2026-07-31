import argparse
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook


def _read_sheet(sheet):
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return None
    headers = list(rows[0])
    if not any(header is not None for header in headers):
        return None
    return pd.DataFrame(rows[1:], columns=headers)


def merge_excelfiles(dir_path, save_path):
    """Merge all workbook sheets in a directory and return the row count."""
    input_dir = Path(dir_path).expanduser()
    output_path = Path(save_path).expanduser()
    if not input_dir.is_dir():
        raise ValueError(f"输入目录不存在：{input_dir}")

    files = [
        path
        for path in sorted(input_dir.glob("*.xlsx"))
        if path.resolve() != output_path.resolve()
    ]
    if not files:
        raise ValueError(f"输入目录中没有可合并的 .xlsx 文件：{input_dir}")

    frames = []
    for path in files:
        workbook = load_workbook(path, read_only=True, data_only=True)
        try:
            for sheet in workbook.worksheets:
                frame = _read_sheet(sheet)
                if frame is not None and not frame.empty:
                    frames.append(frame)
        finally:
            workbook.close()

    if not frames:
        raise ValueError(f"输入文件没有可合并的数据行：{input_dir}")

    merged = pd.concat(frames, axis=0, ignore_index=True, sort=False)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    merged.to_excel(output_path, index=False)
    return len(merged)


def build_parser():
    parser = argparse.ArgumentParser(description="合并目录中的 Excel 工作簿")
    parser.add_argument("--input-dir", required=True, help="Excel 输入目录")
    parser.add_argument("--output", required=True, help="合并后的 Excel 路径")
    return parser


def cli(argv=None):
    args = build_parser().parse_args(argv)
    try:
        count = merge_excelfiles(args.input_dir, args.output)
    except ValueError as exc:
        print(f"合并失败：{exc}")
        return 1
    print(f"已合并 {count} 行：{args.output}")
    return 0


if __name__ == "__main__":
    raise SystemExit(cli())
