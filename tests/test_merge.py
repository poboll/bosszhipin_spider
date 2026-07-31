import tempfile
import unittest
from pathlib import Path

from openpyxl import Workbook, load_workbook

from q import merge_excelfiles


def write_workbook(path, sheets):
    workbook = Workbook()
    workbook.remove(workbook.active)
    for name, rows in sheets:
        sheet = workbook.create_sheet(name)
        for row in rows:
            sheet.append(row)
    workbook.save(path)


class ExcelMergeTests(unittest.TestCase):
    def test_merge_unions_headers_and_skips_output_file(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            directory = Path(temp_dir)
            write_workbook(
                directory / "one.xlsx",
                [("jobs", [["职位", "城市"], ["测试工程师", "福州"]])],
            )
            write_workbook(
                directory / "two.xlsx",
                [("jobs", [["职位", "薪酬"], ["数据分析师", "15-25K"]])],
            )
            output = directory / "merged.xlsx"

            result = merge_excelfiles(directory, output)

            self.assertEqual(result, 2)
            workbook = load_workbook(output, read_only=True, data_only=True)
            rows = list(workbook.active.values)
            workbook.close()
            self.assertEqual(rows[0], ("职位", "城市", "薪酬"))
            self.assertEqual(rows[1], ("测试工程师", "福州", None))
            self.assertEqual(rows[2], ("数据分析师", None, "15-25K"))

    def test_merge_rejects_empty_input_directory(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            with self.assertRaises(ValueError):
                merge_excelfiles(Path(temp_dir), Path(temp_dir) / "merged.xlsx")


if __name__ == "__main__":
    unittest.main()
